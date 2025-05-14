import streamlit as st
import pandas as pd
import pyodbc
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from ERD_Gen import mermaid_to_image
from PIL import Image

# Set page config
st.set_page_config(page_title="Database Metadata Explorer", layout="wide")

# App title
st.title("Database Metadata Explorer")

# Database connection section
st.header("Database Connection")

# Create two columns for DB connection details
col1, col2 = st.columns(2)

with col1:
    db_server = st.text_input("Server", placeholder="e.g., localhost\\SQLEXPRESS")
    db_name = st.text_input("Database Name", placeholder="e.g., AdventureWorks")

with col2:
    db_username = st.text_input("Username (leave empty for Windows Auth)", "")
    db_password = st.text_input("Password (leave empty for Windows Auth)", "", type="password")

# Global variables
if 'conn' not in st.session_state:
    st.session_state.conn = None
if 'tables' not in st.session_state:
    st.session_state.tables = []
if 'selected_table' not in st.session_state:
    st.session_state.selected_table = None
if 'dependencies' not in st.session_state:
    st.session_state.dependencies = {}
if 'similar_tables' not in st.session_state:
    st.session_state.similar_tables = []
if 'search_query' not in st.session_state:
    st.session_state.search_query = ""
if 'filtered_tables' not in st.session_state:
    st.session_state.filtered_tables = []
if 'relationships' not in st.session_state:
    st.session_state.relationships = []

# Function to connect to database
def connect_to_db():
    try:
        # Use Windows authentication if username is empty
        if db_username == "":
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={db_server};DATABASE={db_name};Trusted_Connection=yes;"
        else:
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={db_server};DATABASE={db_name};UID={db_username};PWD={db_password};"
        
        conn = pyodbc.connect(conn_str)
        return conn, "Connected successfully!"
    except Exception as e:
        return None, f"Error connecting to database: {str(e)}"

# Function to get all tables
def get_tables(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE' ORDER BY TABLE_NAME")
    tables = [row[0] for row in cursor.fetchall()]
    cursor.close()
    return tables

# Function to get table columns
def get_table_metadata(conn, table_name):
    cursor = conn.cursor()
    cursor.execute(f"""
    SELECT 
        c.COLUMN_NAME,
        c.DATA_TYPE,
        c.CHARACTER_MAXIMUM_LENGTH,
        c.IS_NULLABLE,
        COLUMNPROPERTY(OBJECT_ID(c.TABLE_SCHEMA + '.' + c.TABLE_NAME), c.COLUMN_NAME, 'IsIdentity') as IS_IDENTITY,
        (SELECT COUNT(*) FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE k
         WHERE k.TABLE_NAME = c.TABLE_NAME AND k.COLUMN_NAME = c.COLUMN_NAME
         AND EXISTS (SELECT 1 FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                    WHERE tc.CONSTRAINT_NAME = k.CONSTRAINT_NAME
                    AND tc.CONSTRAINT_TYPE = 'PRIMARY KEY')) as IS_PRIMARY_KEY
    FROM INFORMATION_SCHEMA.COLUMNS c
    WHERE c.TABLE_NAME = ?
    ORDER BY c.ORDINAL_POSITION
    """, table_name)
    
    columns = []
    for row in cursor.fetchall():
        column_name, data_type, max_length, is_nullable, is_identity, is_primary_key = row
        
        # Format data type with length if applicable
        if max_length is not None and max_length != -1:
            formatted_type = f"{data_type}({max_length})"
        elif max_length == -1:  # MAX types
            formatted_type = f"{data_type}(MAX)"
        else:
            formatted_type = data_type
            
        columns.append({
            "Column Name": column_name,
            "Data Type": formatted_type,
            "Nullable": "YES" if is_nullable == "YES" else "NO",
            "Identity": "YES" if is_identity == 1 else "NO",
            "Primary Key": "YES" if is_primary_key == 1 else "NO"
        })
    
    cursor.close()
    return pd.DataFrame(columns)

# Function to get views
def get_view_metadata(conn, view_name):
    cursor = conn.cursor()
    
    # Get columns in the view
    cursor.execute(f"""
    SELECT 
        c.COLUMN_NAME,
        c.DATA_TYPE,
        c.CHARACTER_MAXIMUM_LENGTH,
        c.IS_NULLABLE
    FROM INFORMATION_SCHEMA.COLUMNS c
    WHERE c.TABLE_NAME = ?
    ORDER BY c.ORDINAL_POSITION
    """, view_name)
    
    columns = []
    for row in cursor.fetchall():
        column_name, data_type, max_length, is_nullable = row
        
        # Format data type with length if applicable
        if max_length is not None and max_length != -1:
            formatted_type = f"{data_type}({max_length})"
        elif max_length == -1:  # MAX types
            formatted_type = f"{data_type}(MAX)"
        else:
            formatted_type = data_type
            
        columns.append({
            "Column Name": column_name,
            "Data Type": formatted_type,
            "Nullable": "YES" if is_nullable == "YES" else "NO"
        })
    
    # Get view definition
    cursor.execute(f"""
    SELECT VIEW_DEFINITION 
    FROM INFORMATION_SCHEMA.VIEWS 
    WHERE TABLE_NAME = ?
    """, view_name)
    
    view_def = cursor.fetchone()
    definition = view_def[0] if view_def else "Definition not available"
    
    cursor.close()
    
    metadata = {
        "columns": pd.DataFrame(columns),
        "definition": definition
    }
    
    return metadata

# Function to get stored procedure metadata
def get_procedure_metadata(conn, proc_name):
    cursor = conn.cursor()
    
    # Get procedure definition
    cursor.execute(f"""
    SELECT ROUTINE_DEFINITION 
    FROM INFORMATION_SCHEMA.ROUTINES 
    WHERE ROUTINE_NAME = ? AND ROUTINE_TYPE = 'PROCEDURE'
    """, proc_name)
    
    proc_def = cursor.fetchone()
    definition = proc_def[0] if proc_def else "Definition not available"
    
    # Get parameters
    cursor.execute(f"""
    SELECT 
        PARAMETER_NAME,
        DATA_TYPE,
        CHARACTER_MAXIMUM_LENGTH,
        PARAMETER_MODE
    FROM INFORMATION_SCHEMA.PARAMETERS
    WHERE SPECIFIC_NAME = ?
    ORDER BY ORDINAL_POSITION
    """, proc_name)
    
    params = []
    for row in cursor.fetchall():
        param_name, data_type, max_length, mode = row
        
        # Format data type with length if applicable
        if max_length is not None and max_length != -1:
            formatted_type = f"{data_type}({max_length})"
        elif max_length == -1:  # MAX types
            formatted_type = f"{data_type}(MAX)"
        else:
            formatted_type = data_type
            
        params.append({
            "Parameter Name": param_name,
            "Data Type": formatted_type,
            "Mode": mode
        })
    
    cursor.close()
    
    metadata = {
        "parameters": pd.DataFrame(params) if params else pd.DataFrame(columns=["Parameter Name", "Data Type", "Mode"]),
        "definition": definition
    }
    
    return metadata

# Function to get function metadata
def get_function_metadata(conn, func_name):
    cursor = conn.cursor()
    
    # Get function definition
    cursor.execute(f"""
    SELECT ROUTINE_DEFINITION 
    FROM INFORMATION_SCHEMA.ROUTINES 
    WHERE ROUTINE_NAME = ? AND ROUTINE_TYPE = 'FUNCTION'
    """, func_name)
    
    func_def = cursor.fetchone()
    definition = func_def[0] if func_def else "Definition not available"
    
    # Get parameters
    cursor.execute(f"""
    SELECT 
        PARAMETER_NAME,
        DATA_TYPE,
        CHARACTER_MAXIMUM_LENGTH,
        PARAMETER_MODE
    FROM INFORMATION_SCHEMA.PARAMETERS
    WHERE SPECIFIC_NAME = ?
    ORDER BY ORDINAL_POSITION
    """, func_name)
    
    params = []
    for row in cursor.fetchall():
        param_name, data_type, max_length, mode = row
        
        # Format data type with length if applicable
        if max_length is not None and max_length != -1:
            formatted_type = f"{data_type}({max_length})"
        elif max_length == -1:  # MAX types
            formatted_type = f"{data_type}(MAX)"
        else:
            formatted_type = data_type
            
        params.append({
            "Parameter Name": param_name,
            "Data Type": formatted_type,
            "Mode": mode
        })
    
    # Get return type
    cursor.execute(f"""
    SELECT 
        DATA_TYPE,
        CHARACTER_MAXIMUM_LENGTH
    FROM INFORMATION_SCHEMA.ROUTINES
    WHERE ROUTINE_NAME = ? AND ROUTINE_TYPE = 'FUNCTION'
    """, func_name)
    
    return_type_row = cursor.fetchone()
    return_type = ""
    if return_type_row:
        data_type, max_length = return_type_row
        
        # Format data type with length if applicable
        if max_length is not None and max_length != -1:
            return_type = f"{data_type}({max_length})"
        elif max_length == -1:  # MAX types
            return_type = f"{data_type}(MAX)"
        else:
            return_type = data_type
    
    cursor.close()
    
    metadata = {
        "parameters": pd.DataFrame(params) if params else pd.DataFrame(columns=["Parameter Name", "Data Type", "Mode"]),
        "return_type": return_type,
        "definition": definition
    }
    
    return metadata

# Function to identify dependencies using the corrected query
def find_dependencies(conn, table_name):
    cursor = conn.cursor()
    dependencies = {
        "tables": [],
        "views": [],
        "procedures": [],
        "functions": []
    }
    
    # Get related tables through foreign keys - UPDATED with correct query
    cursor.execute(f"""
    SELECT DISTINCT
        fk.name AS ForeignKeyName,
        OBJECT_SCHEMA_NAME(fkc.parent_object_id) AS ReferencingSchema,
        tp.name AS ReferencingTable,
        cp.name AS ReferencingColumn,
        OBJECT_SCHEMA_NAME(fkc.referenced_object_id) AS ReferencedSchema,
        tr.name AS ReferencedTable,
        cr.name AS ReferencedColumn
    FROM 
        sys.foreign_keys fk
    JOIN 
        sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
    JOIN 
        sys.tables tp ON fkc.parent_object_id = tp.object_id
    JOIN 
        sys.columns cp ON fkc.parent_object_id = cp.object_id AND fkc.parent_column_id = cp.column_id
    JOIN 
        sys.tables tr ON fkc.referenced_object_id = tr.object_id
    JOIN 
        sys.columns cr ON fkc.referenced_object_id = cr.object_id AND fkc.referenced_column_id = cr.column_id
    WHERE 
        tr.name = ?
    ORDER BY 
        ReferencedTable, ReferencingTable
    """, table_name)
    
    # Store relationship data for ERD generation
    relationships = []
    
    for row in cursor.fetchall():
        fk_name, ref_schema, ref_table, ref_column, refed_schema, refed_table, refed_column = row
        
        # Add to relationships list for Mermaid diagram
        relationships.append({
            "fk_name": fk_name,
            "referencing_schema": ref_schema,
            "referencing_table": ref_table,
            "referencing_column": ref_column,
            "referenced_schema": refed_schema,
            "referenced_table": refed_table,
            "referenced_column": refed_column
        })
        
        # Add to dependencies
        if ref_table != table_name and ref_table not in dependencies["tables"]:
            dependencies["tables"].append(ref_table)
        if refed_table != table_name and refed_table not in dependencies["tables"]:
            dependencies["tables"].append(refed_table)
    
    # Store relationships in session state
    st.session_state.relationships = relationships
    
    # Get views that reference the table
    cursor.execute(f"""
    SELECT DISTINCT
        v.name
    FROM 
        sys.views v
        INNER JOIN sys.sql_expression_dependencies d 
            ON v.object_id = d.referencing_id
        INNER JOIN sys.tables t 
            ON t.object_id = d.referenced_id
    WHERE 
        t.name = ?
    """, table_name)
    
    for row in cursor.fetchall():
        dependencies["views"].append(row[0])
    
    # Get stored procedures that reference the table
    cursor.execute(f"""
    SELECT DISTINCT
        p.name
    FROM 
        sys.procedures p
        INNER JOIN sys.sql_expression_dependencies d 
            ON p.object_id = d.referencing_id
        INNER JOIN sys.tables t 
            ON t.object_id = d.referenced_id
    WHERE 
        t.name = ?
    """, table_name)
    
    for row in cursor.fetchall():
        dependencies["procedures"].append(row[0])
    
    # Get functions that reference the table
    cursor.execute(f"""
    SELECT DISTINCT
        f.name
    FROM 
        sys.objects f
        INNER JOIN sys.sql_expression_dependencies d 
            ON f.object_id = d.referencing_id
        INNER JOIN sys.tables t 
            ON t.object_id = d.referenced_id
    WHERE 
        t.name = ? AND
        f.type IN ('FN', 'IF', 'TF')
    """, table_name)
    
    for row in cursor.fetchall():
        dependencies["functions"].append(row[0])
    
    cursor.close()
    return dependencies

# Function to generate Mermaid ERD
def generate_mermaid_erd(relationships, selected_table):
    mermaid_code = ["erDiagram"]
    
    # Track added relationships to avoid duplicates
    added_relationships = set()
    
    for rel in relationships:
        ref_table = rel["referencing_table"]
        refed_table = rel["referenced_table"]
        
        # Create a unique identifier for this relationship
        rel_key = f"{ref_table}:{refed_table}:{rel['referencing_column']}:{rel['referenced_column']}"
        
        if rel_key not in added_relationships:
            # Add the relationship to the diagram
            # Format: TableA ||--o{ TableB : "Column relationship"
            mermaid_code.append(f'    {refed_table} ||--o{{ {ref_table} : "{rel["referenced_column"]}"')
            added_relationships.add(rel_key)
    
    return "\n".join(mermaid_code)

# Function to find similar tables
def find_similar_tables(tables, selected_table):
    # Extract root name (e.g., "work" from "workorders")
    # Try different patterns to extract meaningful prefixes
    prefixes = []
    
    # Case 1: Split by underscore
    if '_' in selected_table:
        parts = selected_table.split('_')
        # Add the first part as a potential prefix
        if parts[0]:
            prefixes.append(parts[0].lower())
    
    # Case 2: Split by camel case
    # Find boundaries between lowercase and uppercase letters
    camel_parts = re.findall(r'[A-Z]?[a-z]+', selected_table)
    if camel_parts and camel_parts[0]:
        prefixes.append(camel_parts[0].lower())
    
    # Case 3: Just use first 4 characters if they're alphabetic
    if len(selected_table) >= 4 and selected_table[:4].isalpha():
        prefixes.append(selected_table[:4].lower())
    
    # Default: use the table name as is
    prefixes.append(selected_table.lower())
    
    # Find tables that match any of the prefixes
    similar = []
    for table in tables:
        # Skip the selected table itself
        if table.lower() == selected_table.lower():
            continue
            
        # Check if the table starts with any of our prefixes
        table_lower = table.lower()
        for prefix in prefixes:
            if table_lower.startswith(prefix):
                similar.append(table)
                break
                
    return similar

# Function to generate Excel report
def generate_excel_report(conn, selected_table, dependencies, similar_tables):
    # Create workbook
    wb = Workbook()
    
    # Create summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add header
    summary_sheet.append(["Object Type", "Count", "Objects"])
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    all_objects = []
    
    # Add selected table
    table_row = ["Tables", 1 + len(dependencies["tables"]) + len(similar_tables)]
    table_objects = [selected_table] + dependencies["tables"] + similar_tables
    table_row.append(", ".join(table_objects))
    summary_sheet.append(table_row)
    all_objects.extend(table_objects)
    
    # Add views
    if dependencies["views"]:
        views_row = ["Views", len(dependencies["views"]), ", ".join(dependencies["views"])]
        summary_sheet.append(views_row)
        all_objects.extend(dependencies["views"])
    
    # Add procedures
    if dependencies["procedures"]:
        procs_row = ["Stored Procedures", len(dependencies["procedures"]), ", ".join(dependencies["procedures"])]
        summary_sheet.append(procs_row)
        all_objects.extend(dependencies["procedures"])
    
    # Add functions
    if dependencies["functions"]:
        funcs_row = ["Functions", len(dependencies["functions"]), ", ".join(dependencies["functions"])]
        summary_sheet.append(funcs_row)
        all_objects.extend(dependencies["functions"])
    
    # Format summary sheet
    for column in ['A', 'B', 'C']:
        summary_sheet.column_dimensions[column].width = 25 if column != 'C' else 60
    
    # Create sheets for each object
    for obj_name in all_objects:
        # Determine object type
        obj_type = None
        if obj_name == selected_table or obj_name in dependencies["tables"] or obj_name in similar_tables:
            obj_type = "table"
        elif obj_name in dependencies["views"]:
            obj_type = "view"
        elif obj_name in dependencies["procedures"]:
            obj_type = "procedure"
        elif obj_name in dependencies["functions"]:
            obj_type = "function"
        
        if not obj_type:
            continue
        
        # Create sheet for the object
        # Ensure sheet name is valid (max 31 chars, no illegal chars)
        sheet_name = obj_name[:31].replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        obj_sheet = wb.create_sheet(title=sheet_name)
        
        # Add metadata based on object type
        if obj_type == "table":
            metadata = get_table_metadata(conn, obj_name)
            obj_sheet.append(["Table Metadata: " + obj_name])
            obj_sheet.append([])  # Empty row
            
            # Write table metadata
            obj_sheet.append(list(metadata.columns))
            for row in dataframe_to_rows(metadata, index=False, header=False):
                obj_sheet.append(row)
        
        elif obj_type == "view":
            metadata = get_view_metadata(conn, obj_name)
            obj_sheet.append(["View Metadata: " + obj_name])
            obj_sheet.append([])  # Empty row
            
            # Write view columns
            obj_sheet.append(["View Columns:"])
            obj_sheet.append(list(metadata["columns"].columns))
            for row in dataframe_to_rows(metadata["columns"], index=False, header=False):
                obj_sheet.append(row)
            
            # Write view definition
            obj_sheet.append([])  # Empty row
            obj_sheet.append(["View Definition:"])
            obj_sheet.append([metadata["definition"]])
            
        elif obj_type == "procedure":
            metadata = get_procedure_metadata(conn, obj_name)
            obj_sheet.append(["Stored Procedure Metadata: " + obj_name])
            obj_sheet.append([])  # Empty row
            
            # Write procedure parameters
            if not metadata["parameters"].empty:
                obj_sheet.append(["Parameters:"])
                obj_sheet.append(list(metadata["parameters"].columns))
                for row in dataframe_to_rows(metadata["parameters"], index=False, header=False):
                    obj_sheet.append(row)
            
            # Write procedure definition
            obj_sheet.append([])  # Empty row
            obj_sheet.append(["Procedure Definition:"])
            obj_sheet.append([metadata["definition"]])
            
        elif obj_type == "function":
            metadata = get_function_metadata(conn, obj_name)
            obj_sheet.append(["Function Metadata: " + obj_name])
            obj_sheet.append([])  # Empty row
            
            # Write function return type
            obj_sheet.append(["Return Type:"])
            obj_sheet.append([metadata["return_type"]])
            
            # Write function parameters
            if not metadata["parameters"].empty:
                obj_sheet.append([])  # Empty row
                obj_sheet.append(["Parameters:"])
                obj_sheet.append(list(metadata["parameters"].columns))
                for row in dataframe_to_rows(metadata["parameters"], index=False, header=False):
                    obj_sheet.append(row)
            
            # Write function definition
            obj_sheet.append([])  # Empty row
            obj_sheet.append(["Function Definition:"])
            obj_sheet.append([metadata["definition"]])
        
        # Format the sheet
        obj_sheet.column_dimensions['A'].width = 30
        if obj_type in ["table", "view"]:
            for column in ['B', 'C', 'D', 'E']:
                obj_sheet.column_dimensions[column].width = 20
    
    # Save to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# Connect button
if st.button("Connect to Database"):
    conn, message = connect_to_db()
    
    if conn:
        st.session_state.conn = conn
        st.session_state.tables = get_tables(conn)
        st.success(message)
    else:
        st.error(message)

# If connected, show the rest of the app
if st.session_state.conn:
    st.header("Database Exploration")
    
    # Table search functionality
    st.subheader("Search and Select Table")
    search_query = st.text_input("Search for table (partial name search):", 
                                 value=st.session_state.search_query)
    
    # Filter tables based on search query
    if search_query != st.session_state.search_query or not st.session_state.filtered_tables:
        st.session_state.search_query = search_query
        
        if search_query:
            st.session_state.filtered_tables = [table for table in st.session_state.tables 
                                               if search_query.lower() in table.lower()]
        else:
            st.session_state.filtered_tables = st.session_state.tables
    
    # Show filtered tables
    if st.session_state.filtered_tables:
        selected_table = st.selectbox("Select a table:", options=st.session_state.filtered_tables)
        
        if selected_table:
            st.session_state.selected_table = selected_table
            
            # Analyze button
            if st.button("Analyze Database"):
                # Find dependencies and similar tables
                with st.spinner("Analyzing dependencies and similar tables..."):
                    st.session_state.dependencies = find_dependencies(st.session_state.conn, selected_table)
                    st.session_state.similar_tables = find_similar_tables(st.session_state.tables, selected_table)
            
            # Display results if analysis has been performed
            if "dependencies" in st.session_state and st.session_state.dependencies:
                st.subheader("Dependencies")
                
                # Tables
                if st.session_state.dependencies["tables"]:
                    st.write("**Related Tables:**")
                    st.write(", ".join(st.session_state.dependencies["tables"]))
                else:
                    st.write("**Related Tables:** None found")
                
                # Views
                if st.session_state.dependencies["views"]:
                    st.write("**Views:**")
                    st.write(", ".join(st.session_state.dependencies["views"]))
                else:
                    st.write("**Views:** None found")
                
                # Procedures
                if st.session_state.dependencies["procedures"]:
                    st.write("**Stored Procedures:**")
                    st.write(", ".join(st.session_state.dependencies["procedures"]))
                else:
                    st.write("**Stored Procedures:** None found")
                
                # Functions
                if st.session_state.dependencies["functions"]:
                    st.write("**Functions:**")
                    st.write(", ".join(st.session_state.dependencies["functions"]))
                else:
                    st.write("**Functions:** None found")
                
                # Similar tables
                st.subheader("Similar Tables")
                if st.session_state.similar_tables:
                    st.write(", ".join(st.session_state.similar_tables))
                else:
                    st.write("No similar tables found")
                
                # Display Mermaid ERD diagram if relationships exist
                if hasattr(st.session_state, 'relationships') and st.session_state.relationships:
                    st.subheader("Entity Relationship Diagram")
                    mermaid_code = generate_mermaid_erd(st.session_state.relationships, selected_table)
                    st.text_area("Mermaid Code for ERD", mermaid_code, height=300)
                    
                    # Option to download the Mermaid code
                    mermaid_file = BytesIO()
                    mermaid_file.write(mermaid_code.encode('utf-8'))
                    mermaid_file.seek(0)
                    
                    st.download_button(
                        label="Download Mermaid Code",
                        data=mermaid_file,
                        file_name=f"ERD_{selected_table}.mmd",
                        mime="text/plain"
                    )
                    
                    # # Display the ERD using streamlit-mermaid if available
                    # try:
                    #     from streamlit_mermaid import st_mermaid
                    #     st_mermaid(mermaid_code, height=400)
                    # except ImportError:
                    #     st.warning("Install streamlit-mermaid package to render the diagram in the app.")
                    #     st.markdown("### Preview (If supported by your browser):")
                    #     st.markdown(f"```mermaid\n{mermaid_code}\n```")

                    # Convert and show image
                    try:
                        image = mermaid_to_image(mermaid_code, "er_diagram_white.png")
                        image_path = "er_diagram_white.png"
                        image = Image.open(image_path)
                        st.subheader("ERD Diagram")
                        st.image(image, caption="Visualized ERD Diagram", use_column_width=True)

                        # PNG download
                        with open(image_path, "rb") as img_file:
                            st.download_button("Download ERD Diagram (PNG)", data=img_file, file_name="er_diagram.png", mime="image/png")
                    except FileNotFoundError as e:
                        st.error("❌ Failed to generate diagram due to Mermaid syntax error. Please fix the Mermaid code or retry.")
                        st.code(str(e), language="bash")
        
        # Generate Excel report button - only show if analysis has been done
        if "dependencies" in st.session_state and st.session_state.dependencies:
            if st.button("Generate Excel Report"):
                with st.spinner("Generating Excel report..."):
                    excel_file = generate_excel_report(
                        st.session_state.conn,
                        selected_table,
                        st.session_state.dependencies,
                        st.session_state.similar_tables
                    )
                    
                    # Offer download
                    st.download_button(
                        label="Download Excel Report",
                        data=excel_file,
                        file_name=f"DB_Metadata_{selected_table}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("Excel report generated successfully!")

# Cleanup connection when app restarts
def cleanup():
    if st.session_state.conn:
        st.session_state.conn.close()

# Register the cleanup function to be called when the script is terminated
import atexit
atexit.register(cleanup)